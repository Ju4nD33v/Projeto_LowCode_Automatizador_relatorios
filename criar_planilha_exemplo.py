import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import random


VENDEDORES = ["Ana Souza", "Carlos Lima", "Fernanda Reis", "Roberto Neto", "Juliana Melo"]
PRODUTOS = ["Notebook Pro X", "Monitor 4K", "Teclado Mecânico", "Mouse Gamer", "Headset Premium", "Webcam HD"]
REGIOES = ["Sul", "Norte", "Nordeste", "Sudeste", "Centro-Oeste"]


def _gerar_linha() -> dict:
    preco_unit = round(random.uniform(50.0, 3500.0), 2)
    quantidade = random.randint(1, 30)
    return {
        "Data": date(2024, random.randint(1, 12), random.randint(1, 28)).strftime("%d/%m/%Y"),
        "Vendedor": random.choice(VENDEDORES),
        "Produto": random.choice(PRODUTOS),
        "Região": random.choice(REGIOES),
        "Quantidade": quantidade,
        "Preço Unitário (R$)": preco_unit,
        "Total (R$)": round(preco_unit * quantidade, 2),
        "Meta Atingida": random.choice(["Sim", "Não"]),
    }


def criar_planilha_exemplo(caminho: str = "dados_vendas.xlsx") -> None:
    registros = [_gerar_linha() for _ in range(50)]
    df = pd.DataFrame(registros)

    wb = Workbook()
    ws = wb.active
    ws.title = "Vendas 2024"

    HEADER_FILL = PatternFill("solid", fgColor="1A3C5E")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    ALT_FILL = PatternFill("solid", fgColor="EAF0FB")
    BORDER = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    headers = list(df.columns)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    for row_idx, row in df.iterrows():
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)
            cell.border = BORDER
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

    for col_idx, column in enumerate(df.columns, start=1):
        max_len = max(df[column].astype(str).map(len).max(), len(column)) + 4
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 30)

    ws.row_dimensions[1].height = 30

    wb.save(caminho)
    print(f"[OK] Planilha de exemplo criada: '{caminho}'")


if __name__ == "__main__":
    criar_planilha_exemplo()
